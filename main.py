import os
import yaml
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import argparse
from generate_wallet import WalletGenerator, generate_password
from zipfile import BadZipFile

def load_config():
    with open('settings.yaml', 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

def create_excel_if_not_exists(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Wallets"
        headers = [
            "№",
            "MetaMask(ETH) Mnemonic",
            "MetaMask(ETH) Address",
            "MetaMask(ETH) Private Key",
            "SOL Mnemonic", 
            "SOL Address",
            "SOL Private Key",
            "MetaMask Password",
            "Phantom Password",
            "Wallet Password"
        ]
        ws.append(headers)
        # Сделать заголовки жирным шрифтом и установить ширину колонок
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = len(header) + 2
        wb.save(filename)

def append_to_excel(filename, data):
    while True:
        try:
            if os.path.exists(filename):
                book = load_workbook(filename)
                writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
                if 'Wallets' in book.sheetnames:
                    startrow = book['Wallets'].max_row
                else:
                    startrow = 0
                data.to_excel(writer, sheet_name='Wallets', index=False, header=False, startrow=startrow)
                writer.close()
            else:
                create_excel_if_not_exists(filename)
                data.to_excel(filename, index=False, header=True, sheet_name='Wallets')
            break
        except BadZipFile:
            print(f"Error: The file {filename} is corrupted. Creating a new file.")
            create_excel_if_not_exists(filename)
            data.to_excel(filename, index=False, header=True, sheet_name='Wallets')
            break
        except PermissionError:
            input(f"Error: Permission denied for file {filename}. Please close the file if it is open in another program and press Enter to continue.")

def export_data(data, filename, format):
    if format == "xlsx":
        append_to_excel(filename, data)
    elif format == "csv":
        data.to_csv(filename, index=False)
    elif format == "json":
        data.to_json(filename, orient='records', lines=True)
    else:
        raise ValueError(f"Unsupported export format: {format}")

def main():
    config_path = 'settings.yaml'
    config = load_config()
    generator = WalletGenerator(config_path)
    wallets_data = []

    if config['EXPORT_FORMAT'] == "xlsx":
        create_excel_if_not_exists('wallets.xlsx')

    count = int(input("Enter the number of wallets to generate: "))

    # Определить начальный номер строки
    if os.path.exists('wallets.xlsx'):
        book = load_workbook('wallets.xlsx')
        if 'Wallets' in book.sheetnames:
            start_number = book['Wallets'].max_row
        else:
            start_number = 1
    else:
        start_number = 1

    for i in range(count):
        eth_wallet = generator.generate_eth_wallet()
        sol_wallet = generator.generate_sol_wallet()
        if config['RANDOM_PASSWORDS']:
            password = generate_password(config['PASSWORD_LENGTH'])
        else:
            password = config['PASSWORD_WALLET']

        wallet_data = {
            'N': start_number + i,
            'MetaMask(ETH) Mnemonic': eth_wallet['mnemonic'],
            'MetaMask(ETH) Address': eth_wallet['address'],
            'MetaMask(ETH) Private Key': eth_wallet['private_key'],
            'SOL_Mnemonic': sol_wallet['mnemonic'],
            'SOL_Address': sol_wallet['address'],
            'SOL_Private_Key': sol_wallet['private_key'],
            'MetaMask_Password': config['PASSWORD_METAMASK'],
            'Phantom_Password': config['PASSWORD_PHANTOM'],
            'Wallet_Password': password if config['DISPLAY_PASSWORD'] else f"{password[:2]}****{password[-2:]}"
        }
        wallets_data.append(wallet_data)
        print(f"Generated wallet pair {i+1}/{count}")

    df = pd.DataFrame(wallets_data)
    output_path = os.path.abspath(f'wallets.{config["EXPORT_FORMAT"]}')
    export_data(df, output_path, config['EXPORT_FORMAT'])
    print(f"\nDone! Generated {count} wallet pairs and saved to {output_path}")

if __name__ == "__main__":
    main()
